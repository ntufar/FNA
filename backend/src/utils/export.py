"""
Data export utilities for FNA Platform.

Provides functionality to export analysis results to Excel and CSV formats.
"""

import csv
import io
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    Workbook = None

logger = logging.getLogger(__name__)


def export_analysis_to_csv(analyses: List[Dict[str, Any]], filename: Optional[str] = None) -> io.BytesIO:
    """
    Export analysis results to CSV format.
    
    Args:
        analyses: List of analysis dictionaries
        filename: Optional filename (not used, kept for compatibility)
        
    Returns:
        BytesIO: CSV file content as bytes
    """
    output = io.StringIO()
    
    if not analyses:
        output.write("No analysis data available\n")
        return io.BytesIO(output.getvalue().encode('utf-8'))
    
    # Get field names from first analysis
    fieldnames = list(analyses[0].keys())
    
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    
    for analysis in analyses:
        # Flatten nested dictionaries
        flattened = flatten_dict(analysis)
        writer.writerow(flattened)
    
    csv_content = output.getvalue().encode('utf-8')
    return io.BytesIO(csv_content)


def export_analysis_to_excel(analyses: List[Dict[str, Any]], filename: Optional[str] = None) -> io.BytesIO:
    """
    Export analysis results to Excel format.
    
    Args:
        analyses: List of analysis dictionaries
        filename: Optional filename (not used, kept for compatibility)
        
    Returns:
        BytesIO: Excel file content as bytes
        
    Raises:
        ImportError: If openpyxl is not installed
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis Results"
    
    if not analyses:
        ws['A1'] = "No analysis data available"
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    # Get field names from first analysis
    fieldnames = list(flatten_dict(analyses[0]).keys())
    
    # Create header row with styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, fieldname in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=fieldname)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data rows
    for row_idx, analysis in enumerate(analyses, start=2):
        flattened = flatten_dict(analysis)
        for col_idx, fieldname in enumerate(fieldnames, start=1):
            value = flattened.get(fieldname, "")
            # Convert complex types to strings
            if isinstance(value, (dict, list)):
                value = str(value)
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_company_trends_to_csv(company_id: str, trends: List[Dict[str, Any]]) -> io.BytesIO:
    """
    Export company trend data to CSV format.
    
    Args:
        company_id: Company identifier
        trends: List of trend data dictionaries
        
    Returns:
        BytesIO: CSV file content as bytes
    """
    output = io.StringIO()
    
    if not trends:
        output.write(f"No trend data available for company {company_id}\n")
        return io.BytesIO(output.getvalue().encode('utf-8'))
    
    fieldnames = list(trends[0].keys())
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    
    for trend in trends:
        writer.writerow(flatten_dict(trend))
    
    csv_content = output.getvalue().encode('utf-8')
    return io.BytesIO(csv_content)


def export_comparison_to_excel(
    company_id: str,
    comparison_data: Dict[str, Any],
    filename: Optional[str] = None
) -> io.BytesIO:
    """
    Export comparison analysis results to Excel format.
    
    Args:
        company_id: Company identifier
        comparison_data: Comparison analysis data
        filename: Optional filename (not used, kept for compatibility)
        
    Returns:
        BytesIO: Excel file content as bytes
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    wb = Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Add summary data
    summary_data = [
        ["Company ID", company_id],
        ["Comparison Date", datetime.utcnow().isoformat()],
        ["Optimism Delta", comparison_data.get("optimism_delta", "N/A")],
        ["Risk Delta", comparison_data.get("risk_delta", "N/A")],
        ["Uncertainty Delta", comparison_data.get("uncertainty_delta", "N/A")],
        ["Shift Significance", comparison_data.get("shift_significance", "N/A")],
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, start=1):
        ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=2, value=str(value))
    
    # Themes sheet
    if "themes_added" in comparison_data or "themes_removed" in comparison_data:
        ws_themes = wb.create_sheet("Themes")
        
        themes_data = [
            ["Theme", "Status"],
        ]
        
        for theme in comparison_data.get("themes_added", []):
            themes_data.append([theme, "Added"])
        
        for theme in comparison_data.get("themes_removed", []):
            themes_data.append([theme, "Removed"])
        
        for row_idx, row_data in enumerate(themes_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_themes.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def flatten_dict(d: Dict[str, Any], parent_key: str = '', sep: str = '_') -> Dict[str, Any]:
    """
    Flatten a nested dictionary.
    
    Args:
        d: Dictionary to flatten
        parent_key: Parent key prefix
        sep: Separator for nested keys
        
    Returns:
        dict: Flattened dictionary
    """
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            # Convert lists to comma-separated strings
            items.append((new_key, ', '.join(str(item) for item in v)))
        else:
            items.append((new_key, v))
    return dict(items)


def generate_export_filename(prefix: str, format: str = "csv") -> str:
    """
    Generate a filename for exported data.
    
    Args:
        prefix: Filename prefix (e.g., "analysis", "trends")
        format: File format ("csv" or "xlsx")
        
    Returns:
        str: Generated filename
    """
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    extension = "xlsx" if format == "excel" else "csv"
    return f"{prefix}_{timestamp}.{extension}"

